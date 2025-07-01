import os
from io import BytesIO
from typing import Optional, BinaryIO, List
from datetime import datetime

import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter
from copy import copy

from docxtpl import DocxTemplate, InlineImage
from docx.shared import Inches

from app.domain.ports.input_port.report_service import IWordReportUseCase
from app.infrastructure.dto.reports_schema import (
    WordReportDTO, IncidentDTO, ChangeDTO, ServiceRequestDTO,
    AssetDTO, ContactDTO, WorklogDTO
)
from app.api_services.graph_reports_use_case_impl import GraphReportsUseCaseImpl
from app.api_services.tables_reports_use_case_impl import (
    TablesReportsUseCaseImpl,
    build_availability_table_by_month,
)


def fill_val(value, lang):
    if value is None or value == "":
        return "No registra" if lang == "es" else "No data found"
    return str(value)


def filter_incs_closed_only(incidentes: List[IncidentDTO]) -> List[IncidentDTO]:
    closed = []
    for inc in incidentes:
        if inc.status and inc.status.strip().lower() in ["resolved", "closed"]:
            closed.append(inc)
    return closed


def filter_closed_and_open_incidents(incidentes: List[IncidentDTO]):
    closed, opened = [], []
    for inc in incidentes:
        if inc.status:
            st = inc.status.strip().lower()
            if st in ["resolved", "closed"]:
                closed.append(inc)
            elif st == "canceled":
                pass
            else:
                opened.append(inc)
        else:
            opened.append(inc)
    return closed, opened


def filter_closed_and_open_srs(service_requests: List[ServiceRequestDTO]):
    closed, opened = [], []
    for sr in service_requests:
        if sr.status:
            st = sr.status.strip().lower()
            if st in ["resolved", "closed"]:
                closed.append(sr)
            elif st == "canceled":
                pass
            else:
                opened.append(sr)
        else:
            opened.append(sr)
    return closed, opened


def filter_closed_and_open_changes(cambios: List[ChangeDTO]):
    closed, opened = [], []
    for chg in cambios:
        if chg.status:
            st = chg.status.strip().lower()
            if st in ["closed", "review", "completed"]:
                closed.append(chg)
            elif st == "canceled":
                pass
            else:
                opened.append(chg)
        else:
            opened.append(chg)
    return closed, opened


MONTH_NAMES = {
    "es": [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ],
    "en": [
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december"
    ]
}


class UnifiedReportUseCaseImpl(IWordReportUseCase):
    def __init__(self, templates_path: str = "app/utils/templates"):
        self.templates_path = templates_path
        self.graph_reports_use_case = GraphReportsUseCaseImpl()
        self.tables_reports_use_case = TablesReportsUseCaseImpl()

    def generate_report(self, dto: WordReportDTO, template_path: Optional[str] = None) -> BinaryIO:
        if getattr(dto, "report_type", None) == "incidents":
            return self.generate_incidents_report(dto, template_path)
        return self.generate_monthly_report(dto, template_path)

    def generate_monthly_report(self, dto: WordReportDTO, template_path: Optional[str] = None) -> BinaryIO:
        lang = (dto.language or "es").lower()
        closed_incs, open_incs = filter_closed_and_open_incidents(dto.incidentes)
        closed_srs, open_srs = filter_closed_and_open_srs(dto.service_requests)
        closed_chgs, open_chgs = filter_closed_and_open_changes(dto.cambios)
        total_incidentes = len(closed_incs)
        total_service_request = len(closed_srs)
        total_cambios = len(closed_chgs)
        no_data_str = "No registra" if lang == "es" else "No Apply"
        incident_table = self.tables_reports_use_case.build_incident_table(
            closed_incs, lang=lang, no_data_str=no_data_str
        )
        sr_table = self.tables_reports_use_case.build_service_request_table(
            closed_srs, lang=lang, no_data_str=no_data_str
        )
        cambios_table = self.tables_reports_use_case.build_cambios_table(
            closed_chgs, lang=lang, no_data_str=no_data_str
        )
        open_tickets_table = []
        for inc in open_incs:
            open_tickets_table.append({
                "case_type": "INC",
                "case_number": inc.incident_number or no_data_str,
                "ticket_id": inc.ticket_id if inc.ticket_id else no_data_str,
                "status": inc.status or no_data_str,
                "created_at": inc.created_at or no_data_str,
                "description": inc.description or no_data_str
            })
        for sr in open_srs:
            open_tickets_table.append({
                "case_type": "SR",
                "case_number": sr.sr_number or no_data_str,
                "ticket_id": sr.ticket_id if sr.ticket_id else no_data_str,
                "status": sr.status or no_data_str,
                "created_at": sr.created_at.strftime("%d/%m/%Y %H:%M") if sr.created_at else no_data_str,
                "description": sr.symptom or no_data_str
            })
        for chg in open_chgs:
            open_tickets_table.append({
                "case_type": "CHG",
                "case_number": chg.change_number or no_data_str,
                "ticket_id": chg.ticket_id if chg.ticket_id else no_data_str,
                "status": chg.status or no_data_str,
                "created_at": chg.created_at or no_data_str,
                "description": chg.description or no_data_str
            })
        g_proactividad = self.graph_reports_use_case.generate_proactivity_graph(closed_incs, closed_srs, closed_chgs)
        g_top_sedes = self.graph_reports_use_case.generate_top_sedes_graph(closed_incs)
        g_atrib = self.graph_reports_use_case.generate_attributions_graph(closed_incs)
        monthly_avail = build_availability_table_by_month(closed_incs)
        downtime_tables = []
        for (yyyy, mm) in sorted(monthly_avail.keys()):
            data_rows = monthly_avail[(yyyy, mm)]
            if lang == "es":
                mes_arr = MONTH_NAMES["es"]
                mes_name = f"{mes_arr[mm - 1]} de {yyyy}"
            else:
                mes_arr = MONTH_NAMES["en"]
                mes_name = f"{mes_arr[mm - 1]} of {yyyy}"
            downtime_tables.append({
                "periodo": mes_name,
                "rows": data_rows
            })
        if not template_path:
            if lang == "es":
                template_path = os.path.join(self.templates_path, "Informe_Mensual_Estandar.docx")
            else:
                template_path = os.path.join(self.templates_path, "Monthly_Estandard_Report.docx")
        doc = DocxTemplate(template_path)
        months_arr = MONTH_NAMES["en"] if lang == "en" else MONTH_NAMES["es"]
        if dto.start_date:
            start_day_num = dto.start_date.day
            start_month_name = months_arr[dto.start_date.month - 1]
            start_year_str = str(dto.start_date.year)
        else:
            start_day_num = 1
            start_month_name = ""
            start_year_str = ""
        if dto.end_date:
            end_day_num = dto.end_date.day
            end_month_name = months_arr[dto.end_date.month - 1]
            end_year_str = str(dto.end_date.year)
        else:
            end_day_num = 1
            end_month_name = ""
            end_year_str = ""
        ctx = {
            "cust": dto.cust or "No registra",
            "total_incidentes": total_incidentes,
            "total_service_request": total_service_request,
            "total_cambios": total_cambios,
            "incident_table": incident_table,
            "service_request_table": sr_table,
            "cambios_table": cambios_table,
            "open_tickets_table": open_tickets_table,
            "downtime_tables": downtime_tables,
            "start_day_num": start_day_num,
            "end_day_num": end_day_num,
            "start_month_name": start_month_name,
            "start_year_str": start_year_str,
            "end_month_name": end_month_name,
            "end_year_str": end_year_str
        }
        if g_proactividad:
            ctx["grafica_proactividad"] = InlineImage(doc, g_proactividad, width=Inches(6))
        else:
            ctx["grafica_proactividad"] = ""
        if g_top_sedes:
            ctx["grafica_top_sedes"] = InlineImage(doc, g_top_sedes, width=Inches(5))
        else:
            ctx["grafica_top_sedes"] = ""
        if g_atrib:
            ctx["grafica_atribuciones"] = InlineImage(doc, g_atrib, width=Inches(5))
        else:
            ctx["grafica_atribuciones"] = ""
        buffer = BytesIO()
        doc.render(ctx, autoescape=True)
        doc.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_incidents_report(self, dto: WordReportDTO, template_path: Optional[str] = None) -> BinaryIO:
        lang = (dto.language or "es").lower()
        closed_incs = filter_incs_closed_only(dto.incidentes)
        inc_table = self.tables_reports_use_case.build_incident_table(
            closed_incs,
            lang=lang,
            no_data_str="No registra" if lang == "es" else "No Apply"
        )
        monthly_avail = build_availability_table_by_month(closed_incs)
        total_incs = len(closed_incs)
        g_proactividad = self.graph_reports_use_case.generate_proactivity_graph(closed_incs, [], [])
        g_top_sedes = self.graph_reports_use_case.generate_top_sedes_graph(closed_incs)
        g_atrib = self.graph_reports_use_case.generate_attributions_graph(closed_incs)
        if not template_path:
            if lang == "es":
                template_path = os.path.join(self.templates_path, "Informe_Incidencias.docx")
            else:
                template_path = os.path.join(self.templates_path, "Incidents_Report.docx")
        doc = DocxTemplate(template_path)
        if dto.start_date:
            start_day_num = dto.start_date.day
            meslist = MONTH_NAMES["es"] if lang == "es" else MONTH_NAMES["en"]
            s_month_name = meslist[dto.start_date.month - 1]
            s_year_str = str(dto.start_date.year)
        else:
            start_day_num = 1
            s_month_name = ""
            s_year_str = ""
        if dto.end_date:
            end_day_num = dto.end_date.day
            meslist = MONTH_NAMES["es"] if lang == "es" else MONTH_NAMES["en"]
            e_month_name = meslist[dto.end_date.month - 1]
            e_year_str = str(dto.end_date.year)
        else:
            end_day_num = 1
            e_month_name = ""
            e_year_str = ""
        prefix_text = "Disponibilidad para" if lang == "es" else "Availability for"
        downtime_tables = []
        for (yyyy, mm) in sorted(monthly_avail.keys()):
            data_rows = monthly_avail[(yyyy, mm)]
            if lang == "es":
                mesl = MONTH_NAMES["es"]
                mes_str = f"{mesl[mm - 1]} de {yyyy}"
            else:
                mesl = MONTH_NAMES["en"]
                mes_str = f"{mesl[mm - 1]} of {yyyy}"
            downtime_tables.append({
                "periodo": mes_str,
                "rows": data_rows
            })
        context = {
            "cust": dto.cust or "No registra",
            "incident_table": inc_table,
            "downtime_tables": downtime_tables,
            "total_incidentes": total_incs,
            "start_day_num": start_day_num,
            "end_day_num": end_day_num,
            "start_month_name": s_month_name,
            "end_month_name": e_month_name,
            "start_year_str": s_year_str,
            "end_year_str": e_year_str,
            "prefix_text": prefix_text
        }
        if g_proactividad:
            context["grafica_proactividad"] = InlineImage(doc, g_proactividad, width=Inches(5))
        else:
            context["grafica_proactividad"] = ""
        if g_top_sedes:
            context["grafica_top_sedes"] = InlineImage(doc, g_top_sedes, width=Inches(5))
        else:
            context["grafica_top_sedes"] = ""
        if g_atrib:
            context["grafica_atribuciones"] = InlineImage(doc, g_atrib, width=Inches(5))
        else:
            context["grafica_atribuciones"] = ""
        buffer = BytesIO()
        doc.render(context)
        doc.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_saso_excel_report(self, dto: WordReportDTO) -> BinaryIO:
        template_file = os.path.join(self.templates_path, "SASO_Report.xlsx")
        wb = openpyxl.load_workbook(template_file)
        self._fill_customer_table(wb, "Customer Details", "CustomerTable", dto)
        self._fill_asset_table(wb, "Asset Details", "AssetTable", dto)
        self._fill_contact_table(wb, "Contact Details", "ContactTable", dto)
        self._fill_worklog_table(wb, "Worklog Details", "WorklogTable", dto)
        self._fill_incident_table(wb, "Incident Details", "IncidentTable", dto)
        self._fill_sr_table(wb, "ServiceRequest Details", "ServiceRequestTable", dto)
        self._fill_change_table(wb, "Change Details", "ChangeTable", dto)
        out_buffer = BytesIO()
        wb.save(out_buffer)
        out_buffer.seek(0)
        return out_buffer

    def _expand_table_ref(self, ws, table_name: str, start_row: int, count_records: int):
        if table_name not in ws.tables:
            return
        table_obj = ws.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table_obj.ref)
        new_max_row = start_row + count_records - 1
        if new_max_row > max_row:
            max_row = new_max_row
        col_start = get_column_letter(min_col)
        col_end = get_column_letter(max_col)
        table_obj.ref = f"{col_start}{min_row}:{col_end}{max_row}"

    def _fill_customer_table(self, wb, sheet_name: str, table_name: str, dto: WordReportDTO):
        if not dto.customers:
            return
        c = dto.customers[0]
        ws = wb[sheet_name]
        template_row = 3
        row_cells = ws[template_row]
        placeholders = [cell.value for cell in row_cells]
        cell_styles = []
        for cell in row_cells:
            cell_styles.append({
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
                "number_format": cell.number_format,
                "protection": copy(cell.protection)
            })
        for col_idx, ph in enumerate(placeholders, start=1):
            dest = ws.cell(row=template_row, column=col_idx)
            st = cell_styles[col_idx - 1]
            dest.fill = st["fill"]
            dest.font = st["font"]
            dest.alignment = st["alignment"]
            dest.border = st["border"]
            dest.number_format = st["number_format"]
            dest.protection = st["protection"]
            val = "No registra"
            if ph and isinstance(ph, str):
                p = ph.strip()
                if p == "{{ cu.sf_account_id }}":
                    val = c.sf_account_id or "No registra"
                elif p == "{{ cu.name }}":
                    val = c.name or "No registra"
                elif p == "{{ cu.sccd_id }}":
                    val = c.sccd_id or "No registra"
                elif p == "{{ cu.country }}":
                    val = getattr(c, "country", None) or "No registra"
                elif p == "{{ cu.category }}":
                    val = getattr(c, "category", None) or "No registra"
            dest.value = val
        self._expand_table_ref(ws, table_name, template_row, 1)

    def _fill_asset_table(self, wb, sheet_name: str, table_name: str, dto: WordReportDTO):
        ws = wb[sheet_name]
        if not dto.customers or not dto.customers[0].assets:
            return
        assets = dto.customers[0].assets
        template_row = 3
        row_cells = ws[template_row]
        placeholders = [cell.value for cell in row_cells]
        cell_styles = []
        for cell in row_cells:
            cell_styles.append({
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
                "number_format": cell.number_format,
                "protection": copy(cell.protection)
            })
        if len(assets) > 1:
            ws.insert_rows(template_row + 1, amount=len(assets) - 1)
        for i, a in enumerate(assets):
            current_row = template_row + i
            for col_idx, ph in enumerate(placeholders, start=1):
                dest = ws.cell(row=current_row, column=col_idx)
                st = cell_styles[col_idx - 1]
                dest.fill = st["fill"]
                dest.font = st["font"]
                dest.alignment = st["alignment"]
                dest.border = st["border"]
                dest.number_format = st["number_format"]
                dest.protection = st["protection"]
                val = "No registra"
                if ph and isinstance(ph, str):
                    p = ph.strip()
                    if p == "{{ a.sf_asset_id }}":
                        val = a.sf_asset_id or "No registra"
                    elif p == "{{ a.circuit_id }}":
                        val = a.circuit_id or "No registra"
                    elif p == "{{ a.product_family }}":
                        val = a.product_family or "No registra"
                    elif p == "{{ a.product_category }}":
                        val = a.product_category or "No registra"
                    elif p == "{{ a.product_name }}":
                        val = getattr(a, "product_name", None) or "No registra"
                    elif p == "{{ a.status }}":
                        val = getattr(a, "status", None) or "No registra"
                    elif p == "{{ a.location }}":
                        val = a.location or "No registra"
                dest.value = val
        self._expand_table_ref(ws, table_name, template_row, len(assets))

    def _fill_contact_table(self, wb, sheet_name: str, table_name: str, dto: WordReportDTO):
        ws = wb[sheet_name]
        if not dto.customers or not dto.customers[0].contacts:
            return
        contacts = dto.customers[0].contacts
        template_row = 3
        row_cells = ws[template_row]
        placeholders = [cell.value for cell in row_cells]
        cell_styles = []
        for cell in row_cells:
            cell_styles.append({
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
                "number_format": cell.number_format,
                "protection": copy(cell.protection)
            })
        if len(contacts) > 1:
            ws.insert_rows(template_row + 1, amount=len(contacts) - 1)
        for i, c in enumerate(contacts):
            current_row = template_row + i
            for col_idx, ph in enumerate(placeholders, start=1):
                dest = ws.cell(row=current_row, column=col_idx)
                st = cell_styles[col_idx - 1]
                dest.fill = st["fill"]
                dest.font = st["font"]
                dest.alignment = st["alignment"]
                dest.border = st["border"]
                dest.number_format = st["number_format"]
                dest.protection = st["protection"]
                val = "No registra"
                if ph and isinstance(ph, str):
                    p = ph.strip()
                    if p == "{{ c.sf_contact_id }}":
                        val = c.sf_contact_id or "No registra"
                    elif p == "{{ c.name }}":
                        val = c.name or "No registra"
                    elif p == "{{ c.contact_type }}":
                        val = c.contact_type or "No registra"
                    elif p == "{{ c.email }}":
                        val = c.email or "No registra"
                    elif p == "{{ c.phone }}":
                        val = c.phone or "No registra"
                    elif p == "{{ c.account_id }}":
                        val = str(c.account_id) if c.account_id else "No registra"
                dest.value = val
        self._expand_table_ref(ws, table_name, template_row, len(contacts))

    def _fill_worklog_table(self, wb, sheet_name: str, table_name: str, dto: WordReportDTO):
        ws = wb[sheet_name]
        if not dto.customers or not dto.customers[0].worklogs:
            return
        worklogs = dto.customers[0].worklogs
        template_row = 3
        row_cells = ws[template_row]
        placeholders = [cell.value for cell in row_cells]
        cell_styles = []
        for cell in row_cells:
            cell_styles.append({
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
                "number_format": cell.number_format,
                "protection": copy(cell.protection)
            })
        if len(worklogs) > 1:
            ws.insert_rows(template_row + 1, amount=len(worklogs) - 1)
        for i, w in enumerate(worklogs):
            current_row = template_row + i
            for col_idx, ph in enumerate(placeholders, start=1):
                dest = ws.cell(row=current_row, column=col_idx)
                st = cell_styles[col_idx - 1]
                dest.fill = st["fill"]
                dest.font = st["font"]
                dest.alignment = st["alignment"]
                dest.border = st["border"]
                dest.number_format = st["number_format"]
                dest.protection = st["protection"]
                val = "No registra"
                if ph and isinstance(ph, str):
                    p = ph.strip()
                    if p == "{{ w.sf_worklog_id }}":
                        val = w.sf_worklog_id or "No registra"
                    elif p == "{{ w.created_by_name }}":
                        val = w.created_by_name or "No registra"
                    elif p == "{{ w.type_worklog }}":
                        val = w.type_worklog or "No registra"
                    elif p == "{{ w.created_at }}":
                        val = w.created_at.strftime("%Y-%m-%d %H:%M") if w.created_at else "No registra"
                    elif p == "{{ w.description }}":
                        val = w.description or "No registra"
                    elif p == "{{ w.ticket_number }}":
                        val = w.ticket_number or "No registra"
                    elif p == "{{ w.worklog_number }}":
                        val = w.worklog_number or "No registra"
                dest.value = val
        self._expand_table_ref(ws, table_name, template_row, len(worklogs))

    def _fill_incident_table(self, wb, sheet_name: str, table_name: str, dto: WordReportDTO):
        ws = wb[sheet_name]
        incidents = dto.incidentes
        if not incidents:
            return
        template_row = 3
        row_cells = ws[template_row]
        placeholders = [cell.value for cell in row_cells]
        cell_styles = []
        for cell in row_cells:
            cell_styles.append({
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
                "number_format": cell.number_format,
                "protection": copy(cell.protection)
            })
        if len(incidents) > 1:
            ws.insert_rows(template_row + 1, amount=len(incidents) - 1)
        for i, inc in enumerate(incidents):
            asset_id = "No registra"
            asset_loc = "No registra"
            asset_type = "No registra"
            if inc.assets:
                a = inc.assets[0]
                asset_id = a.circuit_id or "No registra"
                asset_loc = a.location or "No registra"
                asset_type = a.product_category or "No registra"
            current_row = template_row + i
            for col_idx, ph in enumerate(placeholders, start=1):
                dest = ws.cell(row=current_row, column=col_idx)
                st = cell_styles[col_idx - 1]
                dest.fill = st["fill"]
                dest.font = st["font"]
                dest.alignment = st["alignment"]
                dest.border = st["border"]
                dest.number_format = st["number_format"]
                dest.protection = st["protection"]
                val = "No registra"
                if ph and isinstance(ph, str):
                    p = ph.strip()
                    if p == "{{ i.subject }}":
                        val = inc.subject or "No registra"
                    elif p == "{{ i.priority }}":
                        val = getattr(inc, "priority", None) or "No registra"
                    elif p == "{{ i.asset }}":
                        val = asset_id
                    elif p == "{{ i.asset_location }}":
                        val = asset_loc
                    elif p == "{{ i.asset_type }}":
                        val = asset_type
                    elif p == "{{ i.sf_incident_id }}":
                        val = inc.sf_incident_id or "No registra"
                    elif p == "{{ i.incident_number }}":
                        val = inc.incident_number or "No registra"
                    elif p == "{{ i.source_incident }}":
                        val = inc.source_incident or "No registra"
                    elif p == "{{ i.reported_at }}":
                        val = inc.reported_at.strftime("%Y-%m-%d %H:%M") if inc.reported_at else "No registra"
                    elif p == "{{ i.affected_at }}":
                        val = inc.affected_at.strftime("%Y-%m-%d %H:%M") if inc.affected_at else "No registra"
                    elif p == "{{ i.resolution_at }}":
                        val = inc.resolution_at.strftime("%Y-%m-%d %H:%M") if inc.resolution_at else "No registra"
                    elif p == "{{ i.status }}":
                        val = inc.status or "No registra"
                    elif p == "{{ i.created_at }}":
                        val = inc.created_at.strftime("%Y-%m-%d %H:%M") if inc.created_at else "No registra"
                    elif p == "{{ i.updated_at }}":
                        val = inc.updated_at.strftime("%Y-%m-%d %H:%M") if inc.updated_at else "No registra"
                    elif p == "{{ i.start_at_dw }}":
                        val = inc.start_at_dw.strftime("%Y-%m-%d %H:%M") if inc.start_at_dw else "No registra"
                    elif p == "{{ i.end_at_dw }}":
                        val = inc.end_at_dw.strftime("%Y-%m-%d %H:%M") if inc.end_at_dw else "No registra"
                    elif p == "{{ i.downtime }}":
                        val = inc.downtime if inc.downtime else 0
                    elif p == "{{ i.is_major }}":
                        val = str(inc.is_major) if inc.is_major is not None else "No registra"
                    elif p == "{{ i.symptom }}":
                        val = inc.symptom or "No registra"
                    elif p == "{{ i.cause }}":
                        val = inc.cause or "No registra"
                    elif p == "{{ i.resolution_summary }}":
                        val = inc.resolution_summary or "No registra"
                    elif p == "{{ i.description }}":
                        val = inc.description or "No registra"
                    elif p == "{{ i.attributed_to }}":
                        val = inc.attributed_to or "No registra"
                    elif p == "{{ i.reason }}":
                        val = inc.reason or "No registra"
                    elif p == "{{ i.type_incident }}":
                        val = inc.type_incident or "No registra"
                    elif p == "{{ i.stop_dw }}":
                        val = inc.stop_dw if inc.stop_dw else 0
                dest.value = val
        self._expand_table_ref(ws, table_name, template_row, len(incidents))

    def _fill_sr_table(self, wb, sheet_name: str, table_name: str, dto: WordReportDTO):
        ws = wb[sheet_name]
        srs = dto.service_requests
        if not srs:
            return
        template_row = 3
        row_cells = ws[template_row]
        placeholders = [cell.value for cell in row_cells]
        cell_styles = []
        for cell in row_cells:
            cell_styles.append({
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
                "number_format": cell.number_format,
                "protection": copy(cell.protection)
            })
        if len(srs) > 1:
            ws.insert_rows(template_row + 1, amount=len(srs) - 1)
        for i, sr in enumerate(srs):
            asset_id = "No registra"
            asset_loc = "No registra"
            asset_type = "No registra"
            if sr.assets:
                a = sr.assets[0]
                asset_id = a.circuit_id or "No registra"
                asset_loc = a.location or "No registra"
                asset_type = a.product_category or "No registra"
            current_row = template_row + i
            for col_idx, ph in enumerate(placeholders, start=1):
                dest = ws.cell(row=current_row, column=col_idx)
                st = cell_styles[col_idx - 1]
                dest.fill = st["fill"]
                dest.font = st["font"]
                dest.alignment = st["alignment"]
                dest.border = st["border"]
                dest.number_format = st["number_format"]
                dest.protection = st["protection"]
                val = "No registra"
                if ph and isinstance(ph, str):
                    p = ph.strip()
                    if p == "{{ s.subject }}":
                        val = sr.subject or "No registra"
                    elif p == "{{ s.priority }}":
                        val = getattr(sr, "priority", None) or "No registra"
                    elif p == "{{ s.asset }}":
                        val = asset_id
                    elif p == "{{ s.asset_location }}":
                        val = asset_loc
                    elif p == "{{ s.asset_type }}":
                        val = asset_type
                    elif p == "{{ s.sf_sr_id }}":
                        val = sr.sf_sr_id or "No registra"
                    elif p == "{{ s.sr_number }}":
                        val = sr.sr_number or "No registra"
                    elif p == "{{ s.status }}":
                        val = sr.status or "No registra"
                    elif p == "{{ s.sr_type }}":
                        val = sr.sr_type or "No registra"
                    elif p == "{{ s.source }}":
                        val = sr.source or "No registra"
                    elif p == "{{ s.symptom }}":
                        val = sr.symptom or "No registra"
                    elif p == "{{ s.solution }}":
                        val = sr.solution or "No registra"
                    elif p == "{{ s.created_at }}":
                        val = sr.created_at.strftime("%Y-%m-%d %H:%M") if sr.created_at else "No registra"
                    elif p == "{{ s.updated_at }}":
                        val = sr.updated_at.strftime("%Y-%m-%d %H:%M") if sr.updated_at else "No registra"
                    elif p == "{{ s.resolved_at }}":
                        val = sr.resolved_at.strftime("%Y-%m-%d %H:%M") if sr.resolved_at else "No registra"
                    elif p == "{{ s.closed_at }}":
                        val = sr.closed_at.strftime("%Y-%m-%d %H:%M") if sr.closed_at else "No registra"
                    elif p == "{{ s.sr_category }}":
                        val = sr.sr_category or "No registra"
                    elif p == "{{ s.sr_type_actions }}":
                        val = sr.sr_type_actions or "No registra"
                dest.value = val
        self._expand_table_ref(ws, table_name, template_row, len(srs))

    def _fill_change_table(self, wb, sheet_name: str, table_name: str, dto: WordReportDTO):
        ws = wb[sheet_name]
        changes = dto.cambios
        if not changes:
            return
        template_row = 3
        row_cells = ws[template_row]
        placeholders = [cell.value for cell in row_cells]
        cell_styles = []
        for cell in row_cells:
            cell_styles.append({
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
                "number_format": cell.number_format,
                "protection": copy(cell.protection)
            })
        if len(changes) > 1:
            ws.insert_rows(template_row + 1, amount=len(changes) - 1)
        for i, ch in enumerate(changes):
            asset_id = "No registra"
            asset_loc = "No registra"
            asset_type = "No registra"
            if ch.assets:
                a = ch.assets[0]
                asset_id = a.circuit_id or "No registra"
                asset_loc = a.location or "No registra"
                asset_type = a.product_category or "No registra"
            current_row = template_row + i
            for col_idx, ph in enumerate(placeholders, start=1):
                dest = ws.cell(row=current_row, column=col_idx)
                st = cell_styles[col_idx - 1]
                dest.fill = st["fill"]
                dest.font = st["font"]
                dest.alignment = st["alignment"]
                dest.border = st["border"]
                dest.number_format = st["number_format"]
                dest.protection = st["protection"]
                val = "No registra"
                if ph and isinstance(ph, str):
                    p = ph.strip()
                    if p == "{{ ch.subject }}":
                        val = ch.subject or "No registra"
                    elif p == "{{ ch.priority }}":
                        val = getattr(ch, "priority", None) or "No registra"
                    elif p == "{{ ch.asset }}":
                        val = asset_id
                    elif p == "{{ ch.asset_location }}":
                        val = asset_loc
                    elif p == "{{ ch.asset_type }}":
                        val = asset_type
                    elif p == "{{ ch.sf_change_id }}":
                        val = ch.sf_change_id or "No registra"
                    elif p == "{{ ch.change_number }}":
                        val = ch.change_number or "No registra"
                    elif p == "{{ ch.status }}":
                        val = ch.status or "No registra"
                    elif p == "{{ ch.type_change }}":
                        val = ch.type_change or "No registra"
                    elif p == "{{ ch.description }}":
                        val = ch.description or "No registra"
                    elif p == "{{ ch.created_at }}":
                        val = ch.created_at.strftime("%Y-%m-%d %H:%M") if ch.created_at else "No registra"
                    elif p == "{{ ch.updated_at }}":
                        val = ch.updated_at.strftime("%Y-%m-%d %H:%M") if ch.updated_at else "No registra"
                    elif p == "{{ ch.result }}":
                        val = ch.result or "No registra"
                    elif p == "{{ ch.type_of_action }}":
                        val = ch.type_of_action or "No registra"
                    elif p == "{{ ch.bussines_reason }}":
                        val = ch.bussines_reason or "No registra"
                    elif p == "{{ ch.urgency }}":
                        val = ch.urgency or "No registra"
                    elif p == "{{ ch.impact }}":
                        val = ch.impact or "No registra"
                    elif p == "{{ ch.risk_level }}":
                        val = ch.risk_level or "No registra"
                    elif p == "{{ ch.failure_probability }}":
                        val = ch.failure_probability or "No registra"
                    elif p == "{{ ch.change_downtime }}":
                        val = ch.change_downtime if ch.change_downtime else 0
                    elif p == "{{ ch.start_at_activity }}":
                        val = ch.start_at_activity.strftime("%Y-%m-%d %H:%M") if ch.start_at_activity else "No registra"
                    elif p == "{{ ch.end_at_activity }}":
                        val = ch.end_at_activity.strftime("%Y-%m-%d %H:%M") if ch.end_at_activity else "No registra"
                    elif p == "{{ ch.bussines_justification }}":
                        val = ch.bussines_justification or "No registra"
                    elif p == "{{ ch.service_impact }}":
                        val = ch.service_impact or "No registra"
                    elif p == "{{ ch.evidence_delivery_at }}":
                        val = ch.evidence_delivery_at.strftime("%Y-%m-%d %H:%M") if ch.evidence_delivery_at else "No registra"
                    elif p == "{{ ch.final_review_at }}":
                        val = ch.final_review_at.strftime("%Y-%m-%d %H:%M") if ch.final_review_at else "No registra"
                    elif p == "{{ ch.cab_assesment }}":
                        val = ch.cab_assesment or "No registra"
                    elif p == "{{ ch.cab_closure }}":
                        val = ch.cab_closure or "No registra"
                    elif p == "{{ ch.category }}":
                        val = ch.category or "No registra"
                    elif p == "{{ ch.client_auth_decision }}":
                        val = ch.client_auth_decision or "No registra"
                dest.value = val
        self._expand_table_ref(ws, table_name, template_row, len(changes))

    def generate_single_case_report(self, case_number: str, dto: WordReportDTO) -> BytesIO:
        lang    = (dto.language or "es").lower()
        no_data = "No registra" if lang == "es" else "No data found"
        inc = dto.incidentes[0] if dto.incidentes else None
        sr  = dto.service_requests[0] if dto.service_requests else None
        ch  = dto.cambios[0] if dto.cambios else None

        inc_ctx = {}
        if inc:
            inc_ctx = {
                "incident_number":    fill_val(inc.incident_number, lang),
                "source_incident":    fill_val(inc.source_incident, lang),
                "status":             fill_val(inc.status, lang),
                "created_at":         fill_val(inc.created_at.strftime("%Y-%m-%d %H:%M") if inc.created_at else "", lang),
                "updated_at":         fill_val(inc.updated_at.strftime("%Y-%m-%d %H:%M") if inc.updated_at else "", lang),
                "is_major":           fill_val(inc.is_major, lang),
                "type_incident":      fill_val(inc.type_incident, lang),
                "description":        fill_val(inc.description, lang),
                "reason":             fill_val(inc.reason, lang),
                "symptom":            fill_val(inc.symptom, lang),
                "cause":              fill_val(inc.cause, lang),
                "resolution_summary": fill_val(inc.resolution_summary, lang),
                "downtime":           fill_val(inc.downtime, lang),
                "stop_dw":            fill_val(inc.stop_dw, lang),
                "attributed_to":      fill_val(inc.attributed_to, lang),
                "start_at_dw":        fill_val(inc.start_at_dw.strftime("%Y-%m-%d %H:%M") if inc.start_at_dw else "", lang),
                "end_at_dw":          fill_val(inc.end_at_dw.strftime("%Y-%m-%d %H:%M") if inc.end_at_dw else "", lang),
                "priority":           fill_val(inc.priority, lang),
                "account_name":       fill_val(getattr(inc, "account_name", None), lang),
            }

        sr_ctx = {}
        if sr:
            sr_ctx = {
                "sr_number":   fill_val(sr.sr_number, lang),
                "source":      fill_val(sr.source, lang),
                "status":      fill_val(sr.status, lang),
                "created_at":  fill_val(sr.created_at.strftime("%Y-%m-%d %H:%M") if sr.created_at else "", lang),
                "updated_at":  fill_val(sr.updated_at.strftime("%Y-%m-%d %H:%M") if sr.updated_at else "", lang),
                "sr_type":     fill_val(sr.sr_type, lang),
                "sr_category": fill_val(sr.sr_category, lang),
                "symptom":     fill_val(sr.symptom, lang),
                "solution":    fill_val(sr.solution, lang),
                "priority":    fill_val(sr.priority, lang),
                "account_name": fill_val(getattr(sr, "account_name", None), lang),
            }

        ch_ctx = {}
        if ch:
            ch_ctx = {
                "change_number":   fill_val(ch.change_number, lang),
                "status":          fill_val(ch.status, lang),
                "created_at":      fill_val(ch.created_at.strftime("%Y-%m-%d %H:%M") if ch.created_at else "", lang),
                "updated_at":      fill_val(ch.updated_at.strftime("%Y-%m-%d %H:%M") if ch.updated_at else "", lang),
                "type_change":     fill_val(ch.type_change, lang),
                "bussines_reason": fill_val(ch.bussines_reason, lang),
                "description":     fill_val(ch.description, lang),
                "result":          fill_val(ch.result, lang),
                "change_downtime": fill_val(ch.change_downtime, lang),
                "end_at_activity": fill_val(ch.end_at_activity.strftime("%Y-%m-%d %H:%M") if ch.end_at_activity else "", lang),
                "priority":        fill_val(ch.priority, lang),
                "account_name":    fill_val(getattr(ch, "account_name", None), lang),
            }

        w_table = []
        if dto.customers and dto.customers[0].worklogs:
            for w in dto.customers[0].worklogs:
                if w.ticket_number == case_number:
                    w_table.append({
                        "created_by_name": fill_val(w.created_by_name, lang),
                        "type_worklog":    fill_val(w.type_worklog, lang),
                        "created_at":      fill_val(w.created_at.strftime("%Y-%m-%d %H:%M") if w.created_at else "", lang),
                        "description":     fill_val(w.description, lang),
                    })

        assets_table = []
        seen = set()
        for obj in (inc, sr, ch):
            if obj and getattr(obj, "assets", None):
                for a in obj.assets:
                    cat = a.product_category or no_data
                    if cat not in seen:
                        seen.add(cat)
                        assets_table.append({
                            "asset":          a.circuit_id or a.product_name or a.sf_asset_id or str(a.asset_id),
                            "asset_location": a.location or no_data,
                            "asset_type":     cat,
                            "account_name":   getattr(a, "account_name", no_data),
                        })
        if not assets_table and dto.customers and dto.customers[0].assets:
            for a in dto.customers[0].assets:
                cat = a.product_category or no_data
                if cat not in seen:
                    seen.add(cat)
                    assets_table.append({
                        "asset":          a.circuit_id or a.product_name or a.sf_asset_id or str(a.asset_id),
                        "asset_location": a.location or no_data,
                        "asset_type":     cat,
                        "account_name":   no_data,
                    })
        if not assets_table:
            assets_table.append({"asset": no_data, "asset_location": no_data, "asset_type": no_data, "account_name": no_data})

        cust_name = (
            dto.cust
            or inc_ctx.get("account_name")
            or sr_ctx.get("account_name")
            or ch_ctx.get("account_name")
            or (assets_table[0].get("account_name") if assets_table else None)
            or (dto.customers[0].name if dto.customers else None)
            or no_data
        )
        if inc and getattr(inc, "is_major", False):
            cust_name = "Multiples clientes afectados" if lang == "es" else "Multiple Customers Impacted"

        template = "Informe_Caso_Individual.docx" if lang == "es" else "Individual_Case_Report.docx"
        doc = DocxTemplate(os.path.join(self.templates_path, template))
        ctx = {
            "cust":           fill_val(cust_name, lang),
            "is_incident":    bool(inc),
            "is_sr":          bool(sr),
            "is_change":      bool(ch),
            "inc":            inc_ctx,
            "sr":             sr_ctx,
            "ch":             ch_ctx,
            "worklogs_table": w_table,
            "assets_table":   assets_table,
        }

        buf = BytesIO()
        doc.render(ctx, autoescape=True)
        doc.save(buf)
        buf.seek(0)
        return buf

    def generate_incident_overview_report(self, dto: WordReportDTO, template_path: Optional[str] = None) -> BytesIO:
        lang = (dto.language or "es").lower()
        inc  = dto.incidentes[0]

        if inc.created_at and inc.end_at_dw:
            delta = (inc.created_at - inc.end_at_dw).total_seconds()
            hrs, mins = divmod(int(delta) // 60, 60)
            duration = f"{hrs:02d}:{mins:02d}"
        else:
            duration = "--:--"

        no_data = "No registra" if lang == "es" else "No data found"

        cust_country = dto.cust_country \
            or (dto.customers[0].country if dto.customers else None) \
            or no_data

        assets_table = []
        if getattr(inc, "assets", None):
            for a in inc.assets:
                assets_table.append({
                    "asset":          a.circuit_id or a.product_name or a.sf_asset_id or str(a.asset_id),
                    "asset_location": a.location or no_data,
                    "asset_type":     a.product_category or no_data,
                })

        unique_types = []
        for entry in assets_table:
            t = entry["asset_type"]
            if t not in unique_types:
                unique_types.append(t)
        cust_assets = ", ".join(unique_types) if unique_types else no_data

        if getattr(inc, "is_major", False):
            cust_display  = "Múltiples clientes afectados" if lang == "es" else "Multiple customers affected"
            type_incident = "Masivo" if lang == "es" else "Massive"
        else:
            cust_display  = dto.cust or no_data
            type_incident = "individual"

        impact_val = getattr(inc, "reason", None) or getattr(inc, "affectation_type", None) or no_data
        if lang == "es":
            trans = {
                "Service Down": "Servicio Caído",
                "Intermittencies and/or Latencies": "Intermitencias o latencias",
                "Service Alarmed": "Servicio Alarmado",
                "Other": "Otros"
            }
            impact_val = trans.get(impact_val, impact_val)
            if getattr(inc, "priority", None):
                prio_map = {"Planning":"Planeación","Low":"Baja","Medium":"Media","High":"Alta","Critical":"Crítica"}
                inc.priority = prio_map.get(inc.priority, inc.priority)

        if not inc.resolution_summary:
            inc.resolution_summary = no_data

        inc_dict = inc.dict()
        inc_dict.update(
            duration=duration,
            type_incident=type_incident,
            afectation_type=impact_val
        )

        path = template_path or os.path.join(self.templates_path, "Incident_Overview_Report.docx")
        doc  = DocxTemplate(path)
        ctx  = {
            "cust":          cust_display,
            "Actual_Time":   dto.report_date or datetime.utcnow().strftime("%Y-%m-%d"),
            "Report_Author": "Customer Service Center",
            "inc":           inc_dict,
            "cust_country":  cust_country,
            "cust_assets":   cust_assets,
            "impact_value":  impact_val,
            "assets_table":  assets_table,
        }

        buf = BytesIO()
        doc.render(ctx, autoescape=True)
        doc.save(buf)
        buf.seek(0)
        return buf
